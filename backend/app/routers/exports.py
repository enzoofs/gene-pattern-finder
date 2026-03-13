import io
import logging
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import AnalysisJob, Collection, JobStatus, CollectionSpecies

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/jobs", tags=["exports"])

# Estilos da planilha (importados lazy no endpoint)


def _style_header(ws, headers: list[str]):
    """Aplica estilo no cabecalho de uma aba."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    """Ajusta largura das colunas ao conteudo."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


@router.get("/{job_id}/export/excel")
async def export_excel(job_id: UUID, db: AsyncSession = Depends(get_db)):
    # Busca job com collection e species
    job = await db.get(AnalysisJob, job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if job.status != JobStatus.done:
        raise HTTPException(400, f"Job not ready: {job.status.value}")

    # Busca especies da colecao
    result = await db.execute(
        select(CollectionSpecies)
        .where(CollectionSpecies.collection_id == job.collection_id)
        .options(
            selectinload(CollectionSpecies.species),
            selectinload(CollectionSpecies.sequence),
        )
    )
    entries = result.scalars().all()

    # Busca colecao pra pegar gene_target
    collection = await db.get(Collection, job.collection_id)

    from openpyxl import Workbook
    wb = Workbook()

    # --- Aba 1: Especies ---
    ws_species = wb.active
    ws_species.title = "Especies"
    headers = ["Nome", "Accession", "Comprimento (bp)", "Gene", "Tipo"]
    _style_header(ws_species, headers)

    for i, entry in enumerate(entries, 2):
        ws_species.cell(row=i, column=1, value=entry.species.name)
        ws_species.cell(row=i, column=2, value=entry.sequence.accession)
        ws_species.cell(row=i, column=3, value=entry.sequence.length)
        ws_species.cell(row=i, column=4, value=collection.gene_target if collection else "")
        ws_species.cell(row=i, column=5, value=entry.sequence.seq_type.value)
    _auto_width(ws_species)

    # --- Aba 2: Conservacao ---
    ws_cons = wb.create_sheet("Conservacao")
    headers_cons = ["Inicio", "Fim", "Comprimento", "Identidade Media (%)"]
    _style_header(ws_cons, headers_cons)

    if job.conservation and "regions" in job.conservation:
        for i, region in enumerate(job.conservation["regions"], 2):
            ws_cons.cell(row=i, column=1, value=region["start"])
            ws_cons.cell(row=i, column=2, value=region["end"])
            ws_cons.cell(row=i, column=3, value=region["length"])
            ws_cons.cell(row=i, column=4, value=round(region["avg_identity"] * 100, 2))
    _auto_width(ws_cons)

    # --- Aba 3: Resumo ---
    ws_summary = wb.create_sheet("Resumo")
    headers_summary = ["Parametro", "Valor"]
    _style_header(ws_summary, headers_summary)

    alignment_len = 0
    if job.alignment:
        # Pega comprimento da primeira sequencia alinhada
        for line in job.alignment.split("\n"):
            if not line.startswith(">") and line.strip():
                alignment_len += len(line.strip())

    conservation_pct = 0.0
    if job.conservation and "conservation_pct" in job.conservation:
        conservation_pct = job.conservation["conservation_pct"]

    summary_data = [
        ("Total de Especies", len(entries)),
        ("Comprimento do Alinhamento", alignment_len),
        ("Conservacao (%)", f"{conservation_pct:.1f}%"),
        ("Modelo da Arvore", job.tree_model or "N/A"),
        ("Gene Alvo", collection.gene_target if collection else "N/A"),
        ("Data da Analise", job.finished_at.strftime("%Y-%m-%d %H:%M") if job.finished_at else "N/A"),
    ]

    for i, (param, valor) in enumerate(summary_data, 2):
        ws_summary.cell(row=i, column=1, value=param)
        ws_summary.cell(row=i, column=2, value=str(valor))
    _auto_width(ws_summary)

    # Gera arquivo em memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gene_pattern_finder_{job_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
